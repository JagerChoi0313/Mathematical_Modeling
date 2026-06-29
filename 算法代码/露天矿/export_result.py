# export_result.py
# 高级版：导出露天矿模型求解结果 Excel
# 包含 Dashboard、模型对比、路线计划、产量验证、品位验证、资源利用、车辆利用、原始参数、建模说明

from pathlib import Path
from collections import defaultdict

import QuestionData as data

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule


# =========================================================
# 1. 输出路径
# =========================================================

output_path = Path(__file__).resolve().parent / "露天矿模型高级求解结果.xlsx"


# =========================================================
# 2. 基础样式
# =========================================================

COLOR_DARK = "1F4E78"
COLOR_BLUE = "D9EAF7"
COLOR_LIGHT_BLUE = "EAF4FB"
COLOR_GREEN = "E2F0D9"
COLOR_YELLOW = "FFF2CC"
COLOR_ORANGE = "FCE4D6"
COLOR_GRAY = "F2F2F2"
COLOR_RED = "F4CCCC"
COLOR_WHITE = "FFFFFF"

thin_side = Side(style="thin", color="BFBFBF")
medium_side = Side(style="medium", color="808080")
border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_medium = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

title_font = Font(name="微软雅黑", size=16, bold=True, color=COLOR_WHITE)
subtitle_font = Font(name="微软雅黑", size=10, color="666666")
header_font = Font(name="微软雅黑", size=10, bold=True, color="000000")
body_font = Font(name="微软雅黑", size=10)
white_bold_font = Font(name="微软雅黑", size=11, bold=True, color=COLOR_WHITE)
kpi_font = Font(name="微软雅黑", size=18, bold=True, color=COLOR_DARK)


# =========================================================
# 3. 模型结果录入
# =========================================================

# 卸点索引：
# 0 = 矿石漏
# 1 = 倒装场I
# 2 = 倒装场II
# 3 = 岩石漏
# 4 = 岩场

model1_routes_raw = [
    {"i": 0, "j": 3, "trucks": 2, "trips": 81},
    {"i": 1, "j": 0, "trucks": 1, "trips": 13},
    {"i": 1, "j": 1, "trucks": 2, "trips": 42},
    {"i": 1, "j": 2, "trucks": 1, "trips": 13},
    {"i": 2, "j": 2, "trucks": 1, "trips": 2},
    {"i": 2, "j": 3, "trucks": 2, "trips": 43},
    {"i": 3, "j": 1, "trucks": 2, "trips": 43},
    {"i": 7, "j": 0, "trucks": 2, "trips": 54},
    {"i": 8, "j": 4, "trucks": 2, "trips": 70},
    {"i": 9, "j": 0, "trucks": 1, "trips": 11},
    {"i": 9, "j": 2, "trucks": 2, "trips": 70},
    {"i": 9, "j": 4, "trucks": 1, "trips": 15},
]

model2_routes_raw = [
    {"i": 0, "j": 3, "trucks": 2, "trips": 81},
    {"i": 1, "j": 0, "trucks": 1, "trips": 15},
    {"i": 1, "j": 1, "trucks": 1, "trips": 37},
    {"i": 1, "j": 2, "trucks": 1, "trips": 16},
    {"i": 1, "j": 3, "trucks": 1, "trips": 28},
    {"i": 3, "j": 1, "trucks": 1, "trips": 37},
    {"i": 3, "j": 3, "trucks": 2, "trips": 51},
    {"i": 4, "j": 0, "trucks": 1, "trips": 22},
    {"i": 4, "j": 1, "trucks": 1, "trips": 35},
    {"i": 4, "j": 4, "trucks": 2, "trips": 39},
    {"i": 7, "j": 0, "trucks": 2, "trips": 52},
    {"i": 7, "j": 2, "trucks": 1, "trips": 32},
    {"i": 8, "j": 4, "trucks": 2, "trips": 76},
    {"i": 9, "j": 2, "trucks": 1, "trips": 47},
    {"i": 9, "j": 4, "trucks": 1, "trips": 45},
]


# =========================================================
# 4. 数据计算函数
# =========================================================

def dump_type(j):
    return "矿石卸点" if j in data.ore_dumps else "岩石卸点"


def material_type(j):
    return "矿石" if j in data.ore_dumps else "岩石"


def enrich_routes(model_name, routes_raw):
    rows = []

    for r in routes_raw:
        i = r["i"]
        j = r["j"]
        trucks = r["trucks"]
        trips = r["trips"]

        amount = trips * data.truck_load
        distance = data.d[i][j]
        ton_km_per_trip = data.ton_km_per_trip[i][j]
        total_ton_km = trips * ton_km_per_trip
        cycle_time = data.cycle_time[i][j]
        max_trips_per_truck = data.max_trips_per_truck[i][j]
        route_capacity = trucks * max_trips_per_truck
        vehicle_util = trips / route_capacity if route_capacity > 0 else 0
        no_wait_limit = int(cycle_time // data.load_time)
        wait_status = "满足" if trucks <= no_wait_limit else "不满足"

        rows.append({
            "模型": model_name,
            "铲位编号": i + 1,
            "铲位": data.shovel_names[i],
            "卸点": data.dump_names[j],
            "卸点类型": dump_type(j),
            "运输物料": material_type(j),
            "距离/km": round(distance, 2),
            "车辆数/辆": trucks,
            "运输趟数/趟": trips,
            "运输量/万吨": round(amount, 4),
            "单趟吨公里": round(ton_km_per_trip, 2),
            "总运量/吨公里": round(total_ton_km, 2),
            "循环时间/min": round(cycle_time, 2),
            "单车最大趟数": max_trips_per_truck,
            "路线总能力/趟": route_capacity,
            "车辆利用率": round(vehicle_util, 4),
            "同路线不等待上限/辆": no_wait_limit,
            "不等待约束": wait_status,
            "矿石品位/%": round(data.grade[i] * 100, 2) if j in data.ore_dumps else "",
            "装车耗时/min": trips * data.load_time,
            "卸车耗时/min": trips * data.unload_time,
        })

    return rows


model1_routes = enrich_routes("模型1_总运量最小", model1_routes_raw)
model2_routes = enrich_routes("模型2_最大产量", model2_routes_raw)
all_routes = model1_routes + model2_routes


def summarize_model(model_name, routes):
    ore_trips = sum(r["运输趟数/趟"] for r in routes if r["运输物料"] == "矿石")
    rock_trips = sum(r["运输趟数/趟"] for r in routes if r["运输物料"] == "岩石")
    total_trips = ore_trips + rock_trips

    ore_amount = ore_trips * data.truck_load
    rock_amount = rock_trips * data.truck_load
    total_amount = total_trips * data.truck_load

    total_transport = sum(r["总运量/吨公里"] for r in routes)
    total_trucks = sum(r["车辆数/辆"] for r in routes)
    selected_shovels = sorted(set(r["铲位"] for r in routes))
    avg_distance = total_transport / (total_trips * data.truck_load_ton) if total_trips > 0 else 0
    transport_intensity = total_transport / total_amount if total_amount > 0 else 0

    return {
        "模型": model_name,
        "矿石运输趟数/趟": ore_trips,
        "岩石运输趟数/趟": rock_trips,
        "总运输趟数/趟": total_trips,
        "矿石产量/万吨": round(ore_amount, 4),
        "岩石产量/万吨": round(rock_amount, 4),
        "总产量/万吨": round(total_amount, 4),
        "总运量/吨公里": round(total_transport, 2),
        "使用卡车数/辆": total_trucks,
        "使用电铲数/台": len(selected_shovels),
        "平均运输距离/km": round(avg_distance, 4),
        "吨公里强度/吨公里每万吨": round(transport_intensity, 2),
        "岩石产量占比": round(rock_amount / total_amount, 4) if total_amount > 0 else 0,
        "电铲布置": "、".join(selected_shovels),
    }


model1_summary = summarize_model("模型1_总运量最小", model1_routes)
model2_summary = summarize_model("模型2_最大产量", model2_routes)


def dump_verification(model_name, routes):
    rows = []

    for j in data.J:
        dump_routes = [r for r in routes if r["卸点"] == data.dump_names[j]]
        trips = sum(r["运输趟数/趟"] for r in dump_routes)
        amount = trips * data.truck_load
        req = data.demand[j]
        excess = amount - req
        satisfaction = amount / req if req > 0 else 0
        unload_util = trips / data.max_trips_per_dump

        rows.append({
            "模型": model_name,
            "卸点": data.dump_names[j],
            "卸点类型": dump_type(j),
            "实际趟数/趟": trips,
            "实际产量/万吨": round(amount, 4),
            "要求产量/万吨": req,
            "超额产量/万吨": round(excess, 4),
            "完成率": round(satisfaction, 4),
            "卸点最大卸车能力/趟": data.max_trips_per_dump,
            "卸点能力利用率": round(unload_util, 4),
            "是否满足": "满足" if amount + 1e-9 >= req else "不满足",
        })

    return rows


dump_rows = dump_verification("模型1_总运量最小", model1_routes) + dump_verification("模型2_最大产量", model2_routes)


def grade_verification(model_name, routes):
    rows = []

    for j in data.ore_dumps:
        dump_routes = [r for r in routes if r["卸点"] == data.dump_names[j]]
        total_trips = sum(r["运输趟数/趟"] for r in dump_routes)

        if total_trips == 0:
            avg_grade = 0
        else:
            weighted_sum = 0
            for r in dump_routes:
                i = r["铲位编号"] - 1
                weighted_sum += data.grade[i] * r["运输趟数/趟"]
            avg_grade = weighted_sum / total_trips

        rows.append({
            "模型": model_name,
            "矿石卸点": data.dump_names[j],
            "矿石运输趟数/趟": total_trips,
            "矿石产量/万吨": round(total_trips * data.truck_load, 4),
            "平均品位/%": round(avg_grade * 100, 4),
            "目标品位/%": 29.5,
            "下限/%": 28.5,
            "上限/%": 30.5,
            "距下限/%": round(avg_grade * 100 - 28.5, 4),
            "距上限/%": round(30.5 - avg_grade * 100, 4),
            "是否满足": "满足" if 28.5 - 1e-9 <= avg_grade * 100 <= 30.5 + 1e-9 else "不满足",
        })

    return rows


grade_rows = grade_verification("模型1_总运量最小", model1_routes) + grade_verification("模型2_最大产量", model2_routes)


def shovel_usage(model_name, routes):
    rows = []

    by_i = defaultdict(list)
    for r in routes:
        i = r["铲位编号"] - 1
        by_i[i].append(r)

    for i in data.I:
        rs = by_i.get(i, [])

        ore_trips = sum(r["运输趟数/趟"] for r in rs if r["运输物料"] == "矿石")
        rock_trips = sum(r["运输趟数/趟"] for r in rs if r["运输物料"] == "岩石")
        total_trips = ore_trips + rock_trips

        used_ore = ore_trips * data.truck_load
        used_rock = rock_trips * data.truck_load

        rows.append({
            "模型": model_name,
            "铲位": data.shovel_names[i],
            "是否布置电铲": "是" if total_trips > 0 else "否",
            "矿石趟数/趟": ore_trips,
            "岩石趟数/趟": rock_trips,
            "总运输趟数/趟": total_trips,
            "电铲最大装车能力/趟": data.max_trips_per_shovel,
            "电铲利用率": round(total_trips / data.max_trips_per_shovel, 4),
            "使用矿石/万吨": round(used_ore, 4),
            "矿石储量/万吨": data.ore[i],
            "矿石使用率": round(used_ore / data.ore[i], 4) if data.ore[i] > 0 else 0,
            "使用岩石/万吨": round(used_rock, 4),
            "岩石储量/万吨": data.rock[i],
            "岩石使用率": round(used_rock / data.rock[i], 4) if data.rock[i] > 0 else 0,
            "矿石品位/%": round(data.grade[i] * 100, 2),
        })

    return rows


shovel_rows = shovel_usage("模型1_总运量最小", model1_routes) + shovel_usage("模型2_最大产量", model2_routes)


def comparison_rows():
    metrics = [
        ("总产量/万吨", model1_summary["总产量/万吨"], model2_summary["总产量/万吨"]),
        ("矿石产量/万吨", model1_summary["矿石产量/万吨"], model2_summary["矿石产量/万吨"]),
        ("岩石产量/万吨", model1_summary["岩石产量/万吨"], model2_summary["岩石产量/万吨"]),
        ("总运输趟数/趟", model1_summary["总运输趟数/趟"], model2_summary["总运输趟数/趟"]),
        ("总运量/吨公里", model1_summary["总运量/吨公里"], model2_summary["总运量/吨公里"]),
        ("使用卡车数/辆", model1_summary["使用卡车数/辆"], model2_summary["使用卡车数/辆"]),
        ("使用电铲数/台", model1_summary["使用电铲数/台"], model2_summary["使用电铲数/台"]),
        ("平均运输距离/km", model1_summary["平均运输距离/km"], model2_summary["平均运输距离/km"]),
        ("岩石产量占比", model1_summary["岩石产量占比"], model2_summary["岩石产量占比"]),
        ("吨公里强度/吨公里每万吨", model1_summary["吨公里强度/吨公里每万吨"], model2_summary["吨公里强度/吨公里每万吨"]),
    ]

    rows = []
    for name, m1, m2 in metrics:
        diff = m2 - m1 if isinstance(m1, (int, float)) and isinstance(m2, (int, float)) else ""
        growth = diff / m1 if isinstance(m1, (int, float)) and m1 != 0 else ""

        rows.append({
            "指标": name,
            "模型1": m1,
            "模型2": m2,
            "模型2-模型1": round(diff, 4) if isinstance(diff, (int, float)) else "",
            "变化率": round(growth, 4) if isinstance(growth, (int, float)) else "",
            "解释": comparison_comment(name, diff, growth),
        })

    return rows


def comparison_comment(name, diff, growth):
    if name == "总产量/万吨":
        return "模型2利用20辆车增产效果明显"
    if name == "岩石产量/万吨":
        return "模型2以岩石产量优先，达到两个岩石卸点能力上限"
    if name == "总运量/吨公里":
        return "模型2产量更高，总运量也相应增加"
    if name == "使用卡车数/辆":
        return "模型2用满20辆车，模型1仅需19辆"
    if name == "吨公里强度/吨公里每万吨":
        return "可用于比较单位产量运输成本"
    return "用于模型结果横向比较"


comparison_data = comparison_rows()


# =========================================================
# 5. Excel 写入工具函数
# =========================================================

def create_sheet(wb, name, title=None, subtitle=None, max_col=10):
    ws = wb.create_sheet(name)

    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        cell = ws.cell(1, 1)
        cell.value = title
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLOR_DARK)
        cell.border = border_medium
        ws.row_dimensions[1].height = 28

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        cell = ws.cell(2, 1)
        cell.value = subtitle
        cell.font = subtitle_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 24

    return ws


def write_table(ws, start_row, start_col, headers, rows, table_name=None):
    # headers
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(start_row, c)
        cell.value = h
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=COLOR_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # rows
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, h in enumerate(headers, start_col):
            cell = ws.cell(r_idx, c_idx)
            cell.value = row.get(h, "")
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_thin

            if isinstance(cell.value, float):
                if "率" in h or "占比" in h or "变化率" in h or "利用率" in h or "完成率" in h:
                    cell.number_format = "0.00%"
                elif "万吨" in h:
                    cell.number_format = "0.0000"
                elif "吨公里" in h or "距离" in h or "时间" in h:
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "0.0000"

    end_row = start_row + len(rows)
    end_col = start_col + len(headers) - 1

    if table_name and len(rows) > 0:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    return end_row, end_col


def adjust_sheet(ws, max_width=32):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A4"


def apply_common_conditional_formats(ws):
    # 对所有“是否满足”列进行颜色标记
    headers = {}
    for col in range(1, ws.max_column + 1):
        headers[ws.cell(3, col).value] = col
        headers[ws.cell(4, col).value] = col

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value == "满足" or value == "是":
                ws.cell(row, col).fill = PatternFill("solid", fgColor=COLOR_GREEN)
            elif value == "不满足" or value == "否":
                ws.cell(row, col).fill = PatternFill("solid", fgColor=COLOR_RED)


def add_data_bar_if_header(ws, header_name, start_row=4):
    col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(start_row, c).value == header_name:
            col = c
            break

    if col is None:
        return

    rng = f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{ws.max_row}"
    ws.conditional_formatting.add(
        rng,
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="63C384",
            showValue=True
        )
    )


# =========================================================
# 6. 创建工作簿
# =========================================================

wb = Workbook()
default_ws = wb.active
wb.remove(default_ws)


# =========================================================
# 7. Sheet 01 Dashboard
# =========================================================

ws = create_sheet(
    wb,
    "01_总览Dashboard",
    "露天矿生产车辆安排模型求解 Dashboard",
    "包含模型1与模型2的关键指标、产量对比、车辆利用与卸点完成情况。",
    max_col=12
)

# KPI 区域
kpis = [
    ("模型1最小总运量", model1_summary["总运量/吨公里"], "吨公里", COLOR_LIGHT_BLUE),
    ("模型1使用卡车", model1_summary["使用卡车数/辆"], "辆", COLOR_LIGHT_BLUE),
    ("模型2总产量", model2_summary["总产量/万吨"], "万吨", COLOR_GREEN),
    ("模型2岩石产量", model2_summary["岩石产量/万吨"], "万吨", COLOR_GREEN),
    ("模型2使用卡车", model2_summary["使用卡车数/辆"], "辆", COLOR_YELLOW),
    ("模型2总运量", model2_summary["总运量/吨公里"], "吨公里", COLOR_YELLOW),
]

positions = [(4, 1), (4, 4), (4, 7), (7, 1), (7, 4), (7, 7)]

for (title, value, unit, fill), (r, c) in zip(kpis, positions):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
    ws.cell(r, c).value = title
    ws.cell(r, c).font = white_bold_font
    ws.cell(r, c).fill = PatternFill("solid", fgColor=COLOR_DARK)
    ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r, c).border = border_medium

    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + 1)
    ws.cell(r + 1, c).value = value
    ws.cell(r + 1, c).font = kpi_font
    ws.cell(r + 1, c).fill = PatternFill("solid", fgColor=fill)
    ws.cell(r + 1, c).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r + 1, c).border = border_medium

    ws.merge_cells(start_row=r + 2, start_column=c, end_row=r + 2, end_column=c + 1)
    ws.cell(r + 2, c).value = unit
    ws.cell(r + 2, c).font = body_font
    ws.cell(r + 2, c).fill = PatternFill("solid", fgColor=COLOR_GRAY)
    ws.cell(r + 2, c).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(r + 2, c).border = border_medium

# Dashboard 数据表：产量对比
production_dashboard = [
    {"指标": "矿石产量/万吨", "模型1": model1_summary["矿石产量/万吨"], "模型2": model2_summary["矿石产量/万吨"]},
    {"指标": "岩石产量/万吨", "模型1": model1_summary["岩石产量/万吨"], "模型2": model2_summary["岩石产量/万吨"]},
    {"指标": "总产量/万吨", "模型1": model1_summary["总产量/万吨"], "模型2": model2_summary["总产量/万吨"]},
]
write_table(ws, 11, 1, ["指标", "模型1", "模型2"], production_dashboard, None)

# Dashboard 数据表：模型2卸点完成情况
model2_dump_dashboard = [
    {"卸点": r["卸点"], "实际产量/万吨": r["实际产量/万吨"], "要求产量/万吨": r["要求产量/万吨"]}
    for r in dump_verification("模型2_最大产量", model2_routes)
]
write_table(ws, 11, 5, ["卸点", "实际产量/万吨", "要求产量/万吨"], model2_dump_dashboard, None)

# 图表1：产量对比
chart1 = BarChart()
chart1.title = "模型1与模型2产量对比"
chart1.y_axis.title = "万吨"
chart1.x_axis.title = "指标"
chart1.height = 7
chart1.width = 12

data_ref = Reference(ws, min_col=2, max_col=3, min_row=11, max_row=14)
cat_ref = Reference(ws, min_col=1, min_row=12, max_row=14)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cat_ref)
chart1.legend.position = "b"
ws.add_chart(chart1, "A17")

# 图表2：模型2卸点实际 vs 要求
chart2 = BarChart()
chart2.title = "模型2各卸点产量完成情况"
chart2.y_axis.title = "万吨"
chart2.x_axis.title = "卸点"
chart2.height = 7
chart2.width = 12

data_ref2 = Reference(ws, min_col=6, max_col=7, min_row=11, max_row=16)
cat_ref2 = Reference(ws, min_col=5, min_row=12, max_row=16)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cat_ref2)
chart2.legend.position = "b"
ws.add_chart(chart2, "F17")

# 关键结论
conclusions = [
    ["关键结论", "说明"],
    ["模型1", "在满足产量与品位要求的前提下，总运量最小，使用19辆卡车。"],
    ["模型2", "利用20辆卡车最大化产量，岩石运输达到320趟，两个岩石卸点卸车能力均用满。"],
    ["品位", "两个模型的矿石卸点平均品位均控制在28.5%至30.5%范围内。"],
    ["资源", "模型2总产量更高，但总运量和运输强度也相应增加。"],
]

for r_idx, row in enumerate(conclusions, 27):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(r_idx, c_idx)
        cell.value = val
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center" if c_idx == 1 else "left", vertical="center", wrap_text=True)
        if r_idx == 27:
            cell.fill = PatternFill("solid", fgColor=COLOR_BLUE)
            cell.font = header_font
        else:
            cell.font = body_font

ws.merge_cells("B28:F28")
ws.merge_cells("B29:F29")
ws.merge_cells("B30:F30")
ws.merge_cells("B31:F31")

adjust_sheet(ws, max_width=28)


# =========================================================
# 8. Sheet 02 模型对比
# =========================================================

ws = create_sheet(
    wb,
    "02_模型对比",
    "模型1与模型2结果对比",
    "模型1侧重运输成本最小；模型2侧重利用现有车辆获得最大产量。",
    max_col=8
)

headers = ["指标", "模型1", "模型2", "模型2-模型1", "变化率", "解释"]
write_table(ws, 4, 1, headers, comparison_data, "CompareTable")
add_data_bar_if_header(ws, "变化率", start_row=4)
adjust_sheet(ws, max_width=36)


# =========================================================
# 9. Sheet 03 模型1生产计划
# =========================================================

ws = create_sheet(
    wb,
    "03_模型1生产计划",
    "模型1生产计划：总运量最小",
    "本表列出模型1的电铲布置、运输路线、车辆数、趟数、运输量和路线能力利用率。",
    max_col=18
)

headers = [
    "模型", "铲位", "卸点", "卸点类型", "运输物料", "距离/km", "车辆数/辆", "运输趟数/趟",
    "运输量/万吨", "单趟吨公里", "总运量/吨公里", "循环时间/min", "单车最大趟数",
    "路线总能力/趟", "车辆利用率", "同路线不等待上限/辆", "不等待约束", "矿石品位/%"
]
write_table(ws, 4, 1, headers, model1_routes, "Model1RouteTable")
add_data_bar_if_header(ws, "车辆利用率", start_row=4)
apply_common_conditional_formats(ws)
adjust_sheet(ws, max_width=24)


# =========================================================
# 10. Sheet 04 模型2生产计划
# =========================================================

ws = create_sheet(
    wb,
    "04_模型2生产计划",
    "模型2生产计划：现有20辆车下最大产量",
    "本表列出模型2的电铲布置、运输路线、车辆数、趟数、运输量和路线能力利用率。",
    max_col=18
)

write_table(ws, 4, 1, headers, model2_routes, "Model2RouteTable")
add_data_bar_if_header(ws, "车辆利用率", start_row=4)
apply_common_conditional_formats(ws)
adjust_sheet(ws, max_width=24)


# =========================================================
# 11. Sheet 05 卸点产量验证
# =========================================================

ws = create_sheet(
    wb,
    "05_卸点产量验证",
    "卸点产量验证",
    "检查各卸点实际产量是否达到题目给定的最低产量要求，并给出卸点能力利用率。",
    max_col=12
)

headers = [
    "模型", "卸点", "卸点类型", "实际趟数/趟", "实际产量/万吨", "要求产量/万吨",
    "超额产量/万吨", "完成率", "卸点最大卸车能力/趟", "卸点能力利用率", "是否满足"
]
write_table(ws, 4, 1, headers, dump_rows, "DumpCheckTable")
add_data_bar_if_header(ws, "完成率", start_row=4)
add_data_bar_if_header(ws, "卸点能力利用率", start_row=4)
apply_common_conditional_formats(ws)
adjust_sheet(ws, max_width=26)


# =========================================================
# 12. Sheet 06 矿石品位验证
# =========================================================

ws = create_sheet(
    wb,
    "06_矿石品位验证",
    "矿石卸点品位验证",
    "检查三个矿石卸点的混合平均品位是否位于28.5%至30.5%的允许区间。",
    max_col=12
)

headers = [
    "模型", "矿石卸点", "矿石运输趟数/趟", "矿石产量/万吨", "平均品位/%",
    "目标品位/%", "下限/%", "上限/%", "距下限/%", "距上限/%", "是否满足"
]
write_table(ws, 4, 1, headers, grade_rows, "GradeCheckTable")
apply_common_conditional_formats(ws)

# 平均品位颜色刻度
for c in range(1, ws.max_column + 1):
    if ws.cell(4, c).value == "平均品位/%":
        rng = f"{get_column_letter(c)}5:{get_column_letter(c)}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="num", start_value=28.5, start_color="F4CCCC",
                mid_type="num", mid_value=29.5, mid_color="FFF2CC",
                end_type="num", end_value=30.5, end_color="D9EAD3"
            )
        )
        break

adjust_sheet(ws, max_width=24)


# =========================================================
# 13. Sheet 07 铲位资源利用
# =========================================================

ws = create_sheet(
    wb,
    "07_铲位资源利用",
    "铲位资源与电铲利用率",
    "展示每个模型下各铲位的电铲布置、矿石/岩石消耗、资源使用率与电铲装车能力利用率。",
    max_col=15
)

headers = [
    "模型", "铲位", "是否布置电铲", "矿石趟数/趟", "岩石趟数/趟", "总运输趟数/趟",
    "电铲最大装车能力/趟", "电铲利用率", "使用矿石/万吨", "矿石储量/万吨", "矿石使用率",
    "使用岩石/万吨", "岩石储量/万吨", "岩石使用率", "矿石品位/%"
]
write_table(ws, 4, 1, headers, shovel_rows, "ShovelUsageTable")
add_data_bar_if_header(ws, "电铲利用率", start_row=4)
add_data_bar_if_header(ws, "矿石使用率", start_row=4)
add_data_bar_if_header(ws, "岩石使用率", start_row=4)
apply_common_conditional_formats(ws)
adjust_sheet(ws, max_width=25)


# =========================================================
# 14. Sheet 08 车辆路线利用
# =========================================================

ws = create_sheet(
    wb,
    "08_车辆路线利用",
    "车辆路线利用率分析",
    "本表合并模型1和模型2的所有运输路线，用于比较路线车辆能力利用率、运输距离和吨公里贡献。",
    max_col=18
)

headers = [
    "模型", "铲位", "卸点", "运输物料", "距离/km", "车辆数/辆", "运输趟数/趟",
    "单车最大趟数", "路线总能力/趟", "车辆利用率", "循环时间/min",
    "运输量/万吨", "单趟吨公里", "总运量/吨公里", "装车耗时/min", "卸车耗时/min",
    "不等待约束"
]
write_table(ws, 4, 1, headers, all_routes, "VehicleRouteTable")
add_data_bar_if_header(ws, "车辆利用率", start_row=4)
apply_common_conditional_formats(ws)
adjust_sheet(ws, max_width=25)


# =========================================================
# 15. Sheet 09 原始数据参数
# =========================================================

ws = create_sheet(
    wb,
    "09_原始数据参数",
    "题目原始数据与参数",
    "本表汇总铲位储量、矿石品位、卸点需求、设备参数与距离矩阵，便于复核。",
    max_col=12
)

# 设备参数
param_rows = [
    {"参数": "卡车载重", "数值": data.truck_load_ton, "单位": "吨/车"},
    {"参数": "卡车载重", "数值": data.truck_load, "单位": "万吨/车"},
    {"参数": "班次时长", "数值": data.shift_time, "单位": "分钟"},
    {"参数": "电铲装车时间", "数值": data.load_time, "单位": "分钟/车"},
    {"参数": "卸点卸车时间", "数值": data.unload_time, "单位": "分钟/车"},
    {"参数": "卡车速度", "数值": data.truck_speed, "单位": "km/h"},
    {"参数": "现有电铲数", "数值": data.num_excavators, "单位": "台"},
    {"参数": "现有卡车数", "数值": data.num_trucks, "单位": "辆"},
    {"参数": "单台电铲最大装车次数", "数值": data.max_trips_per_shovel, "单位": "趟"},
    {"参数": "单个卸点最大卸车次数", "数值": data.max_trips_per_dump, "单位": "趟"},
    {"参数": "品位下限", "数值": data.grade_low * 100, "单位": "%"},
    {"参数": "品位上限", "数值": data.grade_high * 100, "单位": "%"},
]
end_row, _ = write_table(ws, 4, 1, ["参数", "数值", "单位"], param_rows, "ParamTable")

# 铲位数据
shovel_data_rows = []
for i in data.I:
    shovel_data_rows.append({
        "铲位": data.shovel_names[i],
        "矿石量/万吨": data.ore[i],
        "岩石量/万吨": data.rock[i],
        "矿石品位/%": data.grade[i] * 100,
    })

start = end_row + 3
write_table(ws, start, 1, ["铲位", "矿石量/万吨", "岩石量/万吨", "矿石品位/%"], shovel_data_rows, "RawShovelTable")

# 卸点需求
dump_data_rows = []
for j in data.J:
    dump_data_rows.append({
        "卸点": data.dump_names[j],
        "卸点类型": dump_type(j),
        "需求产量/万吨": data.demand[j],
    })

start2 = start + len(shovel_data_rows) + 4
write_table(ws, start2, 1, ["卸点", "卸点类型", "需求产量/万吨"], dump_data_rows, "RawDumpTable")

# 距离矩阵
start3 = start2 + len(dump_data_rows) + 4
distance_headers = ["铲位"] + data.dump_names
distance_rows = []
for i in data.I:
    row = {"铲位": data.shovel_names[i]}
    for j in data.J:
        row[data.dump_names[j]] = data.d[i][j]
    distance_rows.append(row)

write_table(ws, start3, 1, distance_headers, distance_rows, "DistanceTable")
adjust_sheet(ws, max_width=22)


# =========================================================
# 16. Sheet 10 建模说明
# =========================================================

ws = create_sheet(
    wb,
    "10_建模说明",
    "建模说明与结果解释",
    "用于论文撰写时快速回顾模型变量、约束、目标函数和求解结果含义。",
    max_col=8
)

notes = [
    {"模块": "问题性质", "说明": "本题属于带产量、品位、设备能力和车辆能力约束的混合整数规划问题。"},
    {"模块": "主要决策变量", "说明": "x[i,j]表示铲位i到卸点j的运输趟数；y[i]表示铲位i是否布置电铲；n[i,j]表示路线i到j安排的卡车数。"},
    {"模块": "产量约束", "说明": "各卸点实际产量必须不低于题目给定需求；模型2在满足原始需求基础上进一步增产。"},
    {"模块": "品位约束", "说明": "三个矿石卸点的混合平均品位需位于28.5%至30.5%之间，采用线性化后的上下界约束实现。"},
    {"模块": "设备约束", "说明": "最多使用7台电铲；每台电铲每班次最多装车96趟；每个卸点每班次最多卸车160趟。"},
    {"模块": "车辆约束", "说明": "路线运输趟数不能超过该路线车辆数乘以单车最大趟数；总卡车数不超过20辆。"},
    {"模块": "模型1目标", "说明": "在满足各类约束条件下，总运量最小；并进一步要求使用车辆数尽量少。"},
    {"模块": "模型2目标", "说明": "利用现有20辆车获得最大产量，其中岩石产量优先，其次矿石产量，最后考虑总运量较小。"},
    {"模块": "模型1结果", "说明": f"最小总运量为{model1_summary['总运量/吨公里']}吨公里，使用{model1_summary['使用卡车数/辆']}辆卡车，总产量{model1_summary['总产量/万吨']}万吨。"},
    {"模块": "模型2结果", "说明": f"使用20辆卡车，总产量{model2_summary['总产量/万吨']}万吨，其中岩石产量{model2_summary['岩石产量/万吨']}万吨，矿石产量{model2_summary['矿石产量/万吨']}万吨。"},
    {"模块": "结果解读", "说明": "模型1适合运输成本控制场景；模型2适合设备满负荷生产场景。二者目标不同，因此生产计划和总运量不同。"},
]

write_table(ws, 4, 1, ["模块", "说明"], notes, "ModelNotesTable")
adjust_sheet(ws, max_width=60)


# =========================================================
# 17. 全局美化
# =========================================================

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="微软雅黑")
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")

    # 统一页边距和打印设置
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


# =========================================================
# 18. 保存#
# =========================================================

wb.save(output_path)

print("高级版 Excel 导出完成：")
print(output_path)