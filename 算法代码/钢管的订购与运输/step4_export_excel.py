from openpyxl import Workbook

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "运输费用矩阵"

# ==========================
# 写表头
# ==========================

ws.cell(row=1, column=1).value = "钢厂"

for j in range(1, 16):
    ws.cell(row=1, column=j + 1).value = f"A{j}"

# ==========================
# 写矩阵数据
# ==========================

for i in range(1, 8):

    ws.cell(row=i + 1, column=1).value = f"S{i}"

    for j in range(1, 16):

        ws.cell(
            row=i + 1,
            column=j + 1
        ).value = cost_matrix[f"S{i}"][f"A{j}"]

# ==========================
# 保存
# ==========================

filename = "运输费用矩阵.xlsx"

wb.save(filename)

print("=" * 60)
print("Excel导出成功！")
print("文件名：", filename)
print("=" * 60)